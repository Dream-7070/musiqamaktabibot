# ==========================
# services/reports.py
# DIREKTOR UCHUN OYLIK QARZDORLIK HISOBOTI (Excel)
# ==========================


import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


HEADER_FILL = PatternFill(
    start_color="305496",
    end_color="305496",
    fill_type="solid"
)

HEADER_FONT = Font(color="FFFFFF", bold=True)

DEBT_FILL = PatternFill(
    start_color="FCE4E4",
    end_color="FCE4E4",
    fill_type="solid"
)


def _style_header(ws, row=1):

    for cell in ws[row]:

        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws):

    for col in ws.columns:

        length = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=8
        )

        ws.column_dimensions[col[0].column_letter].width = min(length + 3, 45)


def build_debt_report(month, rows):
    """
    rows: [(teacher, department, student, fee, paid_bool, privileged_bool), ...]
    Qaytaradi: (BytesIO, fayl_nomi)
    """

    wb = Workbook()


    # ==========================
    # 1-BET: XULOSA (o'qituvchi bo'yicha)
    # ==========================

    ws1 = wb.active
    ws1.title = "Xulosa"

    ws1.append([
        "O'qituvchi", "Bo'lim", "Jami o'quvchi", "Imtiyozli",
        "Qarzdor soni", "Jami qarz (so'm)"
    ])

    _style_header(ws1)


    summary = {}

    for teacher, dept, student, fee, paid, privileged in rows:

        entry = summary.setdefault(
            teacher,
            {"dept": dept, "total": 0, "free": 0, "unpaid": 0, "debt": 0}
        )

        entry["total"] += 1

        if privileged:
            entry["free"] += 1

        elif not paid:
            entry["unpaid"] += 1
            entry["debt"] += fee


    for teacher, data in sorted(
        summary.items(),
        key=lambda x: x[1]["debt"],
        reverse=True
    ):

        row_index = ws1.max_row + 1

        ws1.append([
            teacher,
            data["dept"],
            data["total"],
            data["free"],
            data["unpaid"],
            data["debt"]
        ])

        if data["debt"] > 0:

            for cell in ws1[row_index]:
                cell.fill = DEBT_FILL


    total_debt = sum(d["debt"] for d in summary.values())
    total_unpaid = sum(d["unpaid"] for d in summary.values())

    ws1.append([])
    total_free = sum(d["free"] for d in summary.values())

    ws1.append(["JAMI", "", "", total_free, total_unpaid, total_debt])

    for cell in ws1[ws1.max_row]:
        cell.font = Font(bold=True)

    _autosize(ws1)


    # ==========================
    # 2-BET: BATAFSIL (o'quvchi bo'yicha)
    # ==========================

    ws2 = wb.create_sheet("Batafsil")

    ws2.append([
        "O'qituvchi", "Bo'lim", "O'quvchi",
        "Oylik badal", "Holat"
    ])

    _style_header(ws2)

    for teacher, dept, student, fee, paid, privileged in rows:

        row_index = ws2.max_row + 1

        if privileged:
            status = "🎖 Imtiyozli"

        elif paid:
            status = "✅ To'langan"

        else:
            status = "❌ Qarzdor"

        ws2.append([
            teacher,
            dept,
            student,
            "-" if privileged else fee,
            status
        ])

        if not paid:

            for cell in ws2[row_index]:
                cell.fill = DEBT_FILL

    _autosize(ws2)


    buffer = io.BytesIO()

    wb.save(buffer)

    buffer.seek(0)

    filename = "hisobot_" + month + ".xlsx"

    return buffer, filename


# ==========================
# O'QUVCHILAR RO'YXATI (Excel)
# ==========================
#
# Telegramda 94 ta o'quvchini matn qilib yuborish o'qib
# bo'lmaydigan uzun ro'yxat hosil qilardi. Endi fayl beriladi.
#
# TARTIB
#   1. Bo'lim - alifbo tartibida
#   2. Shu bo'lim ichida sinf - 1-sinfdan yuqoriga
#   3. Shu sinf ichida o'quvchi - alifbo tartibida
#
# SINF
#   Bazada bir xil emas: "5", "5-sinf", "3 sinf", "2 - sinf",
#   hatto "4 sonf". Shuning uchun raqam matndan ajratib olinadi.
#   Raqami yo'qlari (xato kiritilgan) eng oxirida turadi.
# ==========================


import re


STUDENT_HEADERS = [
    "FISH",
    "Sinfi",
    "Bo'lim",
    "O'qituvchi",
    "ITV raqami",
    "Badal to'lovi miqdori"
]


# imtiyozli o'quvchi belgisi (database.FEE_PRIVILEGED bilan bir xil)
FEE_PRIVILEGED = -1


BAND_FILL = PatternFill(
    start_color="F2F5FA",
    end_color="F2F5FA",
    fill_type="solid"
)

WARN_FILL = PatternFill(
    start_color="FFF3CD",
    end_color="FFF3CD",
    fill_type="solid"
)


# apostrofning har xil ko'rinishlari: o'quvchilar turlicha yozadi
APOSTROPHES = "'\u2018\u2019\u02bb\u02bc\u0060\u00b4"


def uz_sort_key(text):
    """
    Alifbo tartibi uchun kalit.

    Apostroflar olib tashlanadi - shunda "O'ktamov" va "Oktamov"
    yonma-yon turadi (foydalanuvchi ikkalasini bir xil deb biladi).
    Katta-kichik harf farqi ham hisobga olinmaydi.
    """

    clean = (text or "").strip().casefold()

    for mark in APOSTROPHES:
        clean = clean.replace(mark, "")

    return clean


def class_number(text):
    """
    '5', '5-sinf', '2 - sinf', '4 sonf' -> 5, 5, 2, 4
    Raqam topilmasa None (bunday yozuvlar oxirida turadi).
    """

    match = re.search(r"\d+", text or "")

    if not match:
        return None

    value = int(match.group())

    # 7272 kabi xato kiritilgan qiymatlar sinf bo'la olmaydi

    return value if 1 <= value <= 12 else None


def class_label(text):
    """Ko'rsatish uchun bir xil ko'rinish: '5 sinf' -> '5-sinf'."""

    number = class_number(text)

    if number is None:
        return (text or "").strip() or "—"

    return str(number) + "-sinf"


def fee_cell(fee):
    """Excel katakchasi: son bo'lsa son, aks holda izoh."""

    if fee == FEE_PRIVILEGED:
        return "Imtiyozli (bepul)"

    if not fee:
        return "kiritilmagan"

    return int(fee)


def build_students_report(rows):
    """
    rows: [(student, class_name, department, teacher, metrika, fee), ...]
    Qaytaradi: (BytesIO, fayl_nomi)
    """

    ordered = sorted(
        rows,
        key=lambda r: (
            uz_sort_key(r[2]) or "\uffff",              # bo'lim
            class_number(r[1]) if class_number(r[1]) else 99,
            uz_sort_key(r[0])                            # ism-familiya
        )
    )

    wb = Workbook()

    ws = wb.active
    ws.title = "O'quvchilar"

    ws.append(STUDENT_HEADERS)

    _style_header(ws)


    # bo'lim almashganda fon rangi o'zgaradi - uzun ro'yxatda
    # qayerda turganingni ko'rish oson bo'lsin

    previous_department = None
    banded = False

    for student, class_name, department, teacher, metrika, fee in ordered:

        if department != previous_department:
            banded = not banded
            previous_department = department

        ws.append([
            student,
            class_label(class_name),
            department or "—",
            teacher,
            metrika or "—",
            fee_cell(fee)
        ])

        row_index = ws.max_row

        if banded:

            for cell in ws[row_index]:
                cell.fill = BAND_FILL


        # sinfi yoki guvohnomasi kiritilmaganlar ko'zga tashlansin

        if class_number(class_name) is None or not metrika:

            for cell in ws[row_index]:
                cell.fill = WARN_FILL

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = ws.dimensions

    for cell in ws["F"][1:]:
        if isinstance(cell.value, int):
            cell.number_format = "#,##0"

    _autosize(ws)


    buffer = io.BytesIO()

    wb.save(buffer)

    buffer.seek(0)

    return buffer, "oquvchilar_royxati.xlsx"
